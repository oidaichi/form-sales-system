#!/usr/bin/env python3
"""
Flask Webアプリケーション - フォーム営業システム（統合版）
"""
from flask import Flask, render_template, request, jsonify, send_file
from flask_socketio import SocketIO, emit
import threading
import time
import os
from form_sales_system import ExcelProcessor, FormSalesLogger, process_single_company_advanced, EmailSender, process_companies_tab_based_sync

# =============================================================================
# システム設定
# =============================================================================
class Config:
    """システム設定を一元管理"""
    # Excelファイル設定
    SHEET_PRIORITY = ['Sheet1']  # シート検索優先順位
    EXCEL_EXTENSIONS = ['.xlsx', '.xls']
    
    # プレビュー設定
    PREVIEW_COMPANY_COUNT = 10
    
    # アップロード設定
    UPLOAD_FOLDER = 'uploads'

config = Config()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')
socketio = SocketIO(app, cors_allowed_origins="*")

# グローバル変数
current_task = None
task_status = {'running': False, 'progress': 0, 'message': ''}
browser_screenshots = []  # ブラウザスクリーンショット履歴
cdp_session = None  # CDP接続セッション

def screenshot_callback(screenshot_data):
    """スクリーンショットコールバック関数"""
    global browser_screenshots
    
    # 最新の10個のスクリーンショットのみ保持
    browser_screenshots.append(screenshot_data)
    if len(browser_screenshots) > 10:
        browser_screenshots.pop(0)
    
    # WebSocketでクライアントに送信
    socketio.emit('browser_screenshot', screenshot_data)

@app.route('/')
def index():
    """メインページ"""
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Excelファイルアップロード"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'ファイルが選択されていません'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'ファイルが選択されていません'}), 400
        
        if file and file.filename.endswith(tuple(config.EXCEL_EXTENSIONS)):
            # ファイル保存
            upload_path = os.path.join(config.UPLOAD_FOLDER, file.filename)
            os.makedirs(config.UPLOAD_FOLDER, exist_ok=True)
            file.save(upload_path)
            
            # Excel解析 - シート名を自動検出
            processor = ExcelProcessor()
            
            # 利用可能なシート名を取得
            import openpyxl
            workbook = openpyxl.load_workbook(upload_path)
            sheet_names = workbook.sheetnames
            
            # 最適なシート名を選択
            target_sheet = None
            for sheet_name in config.SHEET_PRIORITY + [sheet_names[0] if sheet_names else None]:
                if sheet_name and sheet_name in sheet_names:
                    target_sheet = sheet_name
                    break
            
            if not target_sheet:
                return jsonify({'error': f'利用可能なシートが見つかりません。シート名: {sheet_names}'}), 400
            
            companies = processor.process_excel(upload_path, sheet_name=target_sheet)
            
            # 企業データのプレビューを準備
            preview_data = []
            for company in companies[:config.PREVIEW_COMPANY_COUNT]:
                preview_data.append({
                    'company_name': company.get('company_name', ''),
                    'url': company.get('url', ''),
                    'message': company.get('message', ''),
                    'contact_info': company.get('contact_info', '')
                })
            
            return jsonify({
                'success': True,
                'companies': len(companies),
                'filename': file.filename,
                'preview': preview_data,
                'total_companies': len(companies),
                'sheet_name': target_sheet,
                'available_sheets': sheet_names
            })
        
        return jsonify({'error': '対応していないファイル形式です'}), 400
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/start', methods=['POST'])
def start_processing():
    """営業処理開始"""
    global current_task, task_status
    
    if task_status['running']:
        return jsonify({'error': '既に処理が実行中です'}), 400
    
    try:
        data = request.get_json()
        filename = data.get('filename')
        
        if not filename:
            return jsonify({'error': 'ファイル名が指定されていません'}), 400
        
        # バックグラウンドタスク開始
        task_status = {'running': True, 'progress': 0, 'message': '処理開始中...'}
        current_task = threading.Thread(target=process_companies, args=(filename,))
        current_task.start()
        
        return jsonify({'success': True, 'message': '処理を開始しました'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/start-tabs', methods=['POST'])
def start_tab_processing():
    """タブベース半自動処理を開始"""
    global task_status, current_task
    
    try:
        data = request.get_json()
        if not data or 'filename' not in data:
            return jsonify({'error': 'ファイル名が指定されていません'}), 400
        
        filename = data['filename']
        
        if task_status['running']:
            return jsonify({'error': '他の処理が実行中です'}), 400
        
        task_status = {'running': True, 'progress': 0, 'message': 'タブベース処理を開始しています...'}
        
        current_task = threading.Thread(target=process_companies_with_tabs, args=(filename,))
        current_task.start()
        
        return jsonify({'success': True, 'message': 'タブベース処理を開始しました'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def process_companies_with_tabs(filename):
    """タブベース会社リスト処理"""
    global task_status
    
    try:
        # Excel読み込み
        socketio.emit('status_update', {'message': 'Excelファイルを読み込み中...', 'progress': 10})
        
        processor = ExcelProcessor()
        
        # シート名を自動検出
        upload_path = f'{config.UPLOAD_FOLDER}/{filename}'
        import openpyxl
        workbook = openpyxl.load_workbook(upload_path)
        sheet_names = workbook.sheetnames
        
        target_sheet = None
        for sheet_name in config.SHEET_PRIORITY + [sheet_names[0] if sheet_names else None]:
            if sheet_name and sheet_name in sheet_names:
                target_sheet = sheet_name
                break
        
        companies = processor.process_excel(upload_path, sheet_name=target_sheet)
        
        if not companies:
            socketio.emit('status_update', {'message': '処理する企業データがありません', 'progress': 100})
            task_status['running'] = False
            return
        
        # タブベース処理実行
        socketio.emit('status_update', {'message': f'GUI Chrome起動中... {len(companies)}社を処理します', 'progress': 20})
        
        result = process_companies_tab_based_sync(companies)
        
        if result['success']:
            # 処理完了
            socketio.emit('status_update', {
                'message': f'✅ {result["message"]}', 
                'progress': 100
            })
            
            # 詳細結果をログ出力
            print(f"🎯 タブベース処理完了:")
            print(f"  - 総処理数: {result['total_processed']}社")
            print(f"  - 自動送信成功: {result['completed_count']}社")
            print(f"  - 人間作業待ち: {result['pending_manual_count']}社")
            print(f"  - アクティブタブ: {result['active_tabs_count']}個")
            
            if result['active_tabs_info']:
                print("🖥️ 残っているタブ（人間作業待ち）:")
                for tab_info in result['active_tabs_info']:
                    print(f"  - [{tab_info['company_index']}] {tab_info['company_name']}: {tab_info['status']}")
            
        else:
            socketio.emit('status_update', {
                'message': f'❌ 処理エラー: {result["error"]}', 
                'progress': 100
            })
        
    except Exception as e:
        print(f"❌ タブベース処理エラー: {str(e)}")
        socketio.emit('status_update', {
            'message': f'❌ エラー: {str(e)}', 
            'progress': 100
        })
        
    finally:
        task_status['running'] = False

def process_companies(filename):
    """会社リスト処理"""
    global task_status
    
    try:
        # Excel読み込み
        socketio.emit('status_update', {'message': 'Excelファイルを読み込み中...', 'progress': 10})
        
        processor = ExcelProcessor()
        
        # シート名を自動検出（アップロード時と同じロジック）
        upload_path = f'{config.UPLOAD_FOLDER}/{filename}'
        import openpyxl
        workbook = openpyxl.load_workbook(upload_path)
        sheet_names = workbook.sheetnames
        
        target_sheet = None
        for sheet_name in config.SHEET_PRIORITY + [sheet_names[0] if sheet_names else None]:
            if sheet_name and sheet_name in sheet_names:
                target_sheet = sheet_name
                break
        
        companies = processor.process_excel(upload_path, sheet_name=target_sheet)
        
        total_companies = len(companies)
        results = []  # 処理結果を保存
        
        # 各会社を処理
        for i, company in enumerate(companies):
            if not task_status['running']:
                print(f"🛑 処理停止要求: {i+1}/{total_companies}社で中断")
                break
            
            progress = int((i / total_companies) * 80) + 10  # 10-90%
            message = f"{company['company_name']} を処理中... ({i+1}/{total_companies})"
            
            socketio.emit('status_update', {'message': message, 'progress': progress})
            print(f"🏢 [{i+1}/{total_companies}] 開始: {company['company_name']}")
            
            # フォーム処理実行
            try:
                result = process_single_company(company)
                result['row_index'] = company['index']  # Excel行番号を追加
                results.append(result)
                
                # 処理結果のログ出力
                status_icon = "✅" if result.get('submit_success') else "⚠️" if result.get('form_found') else "❌"
                print(f"{status_icon} [{i+1}/{total_companies}] 完了: {company['company_name']} - {result.get('status_message', '')}")
                
            except Exception as company_error:
                print(f"❌ [{i+1}/{total_companies}] エラー: {company['company_name']} - {company_error}")
                # エラーでも結果を記録
                error_result = {
                    'company_name': company.get('company_name', ''),
                    'url': company.get('url', ''),
                    'form_found': False,
                    'submit_attempted': False,
                    'submit_success': False,
                    'status_message': f'処理エラー: {str(company_error)}',
                    'errors': [str(company_error)],
                    'row_index': company['index']
                }
                results.append(error_result)
            
            # 進捗更新
            progress_message = f"完了: {i+1}/{total_companies}社"
            socketio.emit('status_update', {'message': progress_message, 'progress': progress})
            
            time.sleep(1)  # レート制限（短縮）
        
        # 結果をExcelに保存
        if results:
            socketio.emit('status_update', {'message': '結果をExcelファイルに保存中...', 'progress': 95})
            
            output_path = processor.save_results_to_excel(upload_path, target_sheet, results)
            if output_path:
                socketio.emit('status_update', {'message': f'結果保存完了: {os.path.basename(output_path)}', 'progress': 100})
            else:
                socketio.emit('status_update', {'message': '結果保存でエラーが発生しました', 'progress': 100})
        
        task_status = {'running': False, 'progress': 100, 'message': '処理完了'}
        socketio.emit('status_update', {'message': '全ての処理が完了しました', 'progress': 100})
        
    except Exception as e:
        task_status = {'running': False, 'progress': 0, 'message': f'エラー: {str(e)}'}
        socketio.emit('status_update', {'message': f'エラーが発生しました: {str(e)}', 'progress': 0})

def process_single_company(company):
    """単一会社の処理（統合システム使用）"""
    return process_single_company_advanced(company, enable_screenshots=True, screenshot_callback=screenshot_callback)

@app.route('/api/stop', methods=['POST'])
def stop_processing():
    """処理停止"""
    global task_status
    task_status['running'] = False
    return jsonify({'success': True, 'message': '処理を停止しました'})

@app.route('/api/status')
def get_status():
    """ステータス取得"""
    return jsonify(task_status)

@app.route('/api/logs')
def get_logs():
    """ログ取得"""
    try:
        logger = FormSalesLogger()
        logs = logger.get_recent_logs(100)
        return jsonify({'logs': logs})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/screenshots')
def get_screenshots():
    """ブラウザスクリーンショット履歴取得"""
    global browser_screenshots
    return jsonify({'screenshots': browser_screenshots})

@app.route('/api/cdp-info')
def get_cdp_info():
    """CDP接続情報取得（動的ポート対応）"""
    from form_sales_system import PortManager
    
    crawler_port = PortManager.get_assigned_port('crawler') or 9223
    filler_port = PortManager.get_assigned_port('form_filler') or 9222
    
    return jsonify({
        'crawler_cdp': f'http://localhost:{crawler_port}',
        'filler_cdp': f'http://localhost:{filler_port}',
        'available': True,
        'ports': {
            'crawler': crawler_port,
            'filler': filler_port
        }
    })

@app.route('/browser-embed')
def browser_embed():
    """ブラウザ埋め込み画面"""
    return render_template('browser_embed.html')

@socketio.on('connect')
def handle_connect():
    """WebSocket接続"""
    emit('connected', {'message': '接続しました'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    socketio.run(app, host='0.0.0.0', port=port, debug=False, allow_unsafe_werkzeug=True)
